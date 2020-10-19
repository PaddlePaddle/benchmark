
from openpyxl import Workbook
from openpyxl import load_workbook
import re

sheet = {'>100x': [], '10x~100x':[], '1x~10x':[], '30%~1x':[], '5%~30%':[], 'gpu_backward':[]}
wb = Workbook()
# ws = wb.active
del wb['Sheet']
# ws = wb["Sheet"]
# wb.remove[ws]
for i in sheet.keys():
   ws = wb.create_sheet()
   ws.title = i
wb.save('gpu_backward.xlsx')

src_wb = load_workbook('/benchmark/api/deploy/op_summary_GPU_2.0_1.8_1016.xlsx')
dest_wb = load_workbook('gpu_backward.xlsx')

src_sheet = src_wb['gpu_backward']
# dest_sheet = dest_wb['gpu_backward']

rows = src_sheet.max_row
# gpu_backward_list = []
# sheet = {'gpu_backward': gpu_backward_list}
for i in sheet.keys():
#  dest_sheet = dest_wb[i]
  sheet[i].append(1)
#  for j in range(1, src_sheet.max_column):
#        dest_sheet.column_dimensions[j].width = 40

column_data = []
sub_str = '%'
for i in range(2, rows + 1):
      cell_value = src_sheet.cell(row=i, column=5).value
      if cell_value != '--' and '差于' in cell_value:
         sheet['gpu_backward'].append(i)
         ration=re.findall(r'[(](.*?)[)]', cell_value)
#         ration = cell_value.substring(cell_value.indexOf("(")+1,st1.indexOf(")"))
         ration = ''.join(ration)
         flag = sub_str in ration
         if flag:
             ration_value = ration.strip('%')
             ration_value = float(ration_value)/100
         else: 
             ration_value = ration.strip('x')
             print(ration_value)
         ration_value = float(ration_value)
         if ration_value > 0.05 and ration_value <= 0.3:
              sheet['5%~30%'].append(i)
         elif ration_value > 0.3 and ration_value <= 1:
              sheet['30%~1x'].append(i)
         elif ration_value > 1 and ration_value <= 10:
              sheet['1x~10x'].append(i)
         elif ration_value > 10 and ration_value <= 100:
              sheet['10x~100x'].append(i)
         elif ration_value > 100:
              sheet['>100x'].append(i)
#         for j in range(1, src_sheet.max_column+1):
#           dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value
      column_data.append(cell_value)
print(sheet['gpu_backward'])
print(len(sheet['gpu_backward']))

# for i in range(1, src_sheet.max_row+1):
# row_value = sheet['gpu_backward']
for ws in sheet.keys():
   row_value = sheet[ws]
   dest_sheet = dest_wb[ws]
   for i in range(len(row_value)):
        for j in range(1, src_sheet.max_column):
            dest_sheet.cell(row=i+1, column=j).value = src_sheet.cell(row=row_value[i], column=j).value


dest_wb.save('gpu_backward.xlsx')
